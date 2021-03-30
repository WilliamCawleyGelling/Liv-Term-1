''' 
William Cawley Gelling 
201077658
Assignment 5 Cards 
'''

import openpyxl

class Card(): 
    '''
    Card object which stores information on the card 
    ''' 
    
    #To check the incoming types against 
    types = ["Magi", "Water", "Fire", "Earth", "Air", "Astral"]

    def __init__(self, theName, theType, theHP, isShiny, theMoves):
        
        try: 
            if (isinstance(theName, str) and theName.replace(' ','').isalpha() ): 
                self.name = theName
            else: 
                raise Exception ("Name has to be a string containing only letters.")
        except Exception as exp: 
            print('Error : ', exp)

        try: 
            if (theType.capitalize() in Card.types): 
                self.type = theType.capitalize()
            else: 
                raise Exception ("This is not an aloud type")
        except Exception as exp: 
            print('Error : ', exp)

        try: 
            if (theHP > 0): 
                self.hp = theHP 
            else :
                raise Exception ("HP has to be a positive number ")
        except Exception as exp: 
            print('Error : ', exp)

        try: 
            if (isShiny == 0 or isShiny == 1):
                self.shiny = isShiny
            else: 
                raise Exception ("IsShiny can be either 0 or 1")
        except Exception as exp: 
            print('Error : ', exp)
        self.moves = {} 
        try: 
            if (isinstance(theMoves, list) and len(theMoves)>0 and len(theMoves) <= 5): 
                
                for i in theMoves: 

                    if (isinstance(i, tuple) and len(i) == 2) : 

                        if (isinstance(i[0], str ) and i[0].strip(' ') != ''): 
                            #does this deal with numbers being switched into a name 

                            if (i[1]>0 and isinstance(i[1], int)): 
                                #this may always not work as it sees i[1] as a sting, could cast it as an int and raise a type error if it is not. 
                                self.moves.setdefault(i[0],i[1])
                            elif (i[1] == None):
                                continue
                            else: 
                                raise Exception ('The move ', i[0], 'has to have a positive interger damage factor')

                        elif (i[0] == None):
                            continue
                        else: 
                            raise Exception ('The move ', i[0], 'has to be a non empty string ')

                    else : 
                        raise Exception ('theMoves has to be a list of touples with 2 entries of form (Name, Damage Factor)')

            else : 
                raise Exception ('theMoves has to be contain between one and five tuples representing (Name, Damage Factor )')
            
            if len(self.moves) <= 0 : 
                raise Exception ('There are no moves!')

        except Exception as exp: 
            self.moves = None
            print('Error : ', exp)

        try: 
            if isinstance(self.moves,dict) and len(self.moves)>0 : 
                self.totalDam = 0
                for key in self.moves: 
                    self.totalDam += self.moves[key]
                self.aveDam = self.totalDam/len(self.moves)

            else: 
                raise Exception ('Moves have to be set to take average')
        
        except Exception as exp: 
            self.moves = None
            print('Error : ', exp)
            

    def __str__(self):
        return "\nName: " + self.name + "\nType: " + self.type + "\nMaximum Health Points: " + str(self.hp) + "\nShiny Status: " + str(self.shiny) + '\n' + self.strRepOfMoves() + '\n'


    def strRepOfMoves(self): 
        """
        creates a sting representation of the moves

        Parameters:
            Nothing
        Returns:
            The string representation of the moves 
        """

        strMoves = 'Moves: ' 
        for key in self.moves : 
            strMoves  += '\n' + str(key) + ' : ' + str(self.moves[key])
        return strMoves


class Deck(): 

    def __init__(self): 
        self.deck = [] 
        self.deckLength = 0 
        self.noOfShiny = 0
        self.totalAveDam = 0

    def __str__(self): 
        return 'The total number of cards in the deck is: '+ str(self.deckLength) + '\nThe total number of shiney cards is: ' + str(self.noOfShiny) + '\nThe average damage of the deck is: ' + str(self.getAverageDamage())

    def inputFromFile(self,fileName):
        """
        Inputs a xlsx file matching the row format into cards and inserts them into the deck.

        Parameters:
            fileName - the file name and path if its not in the same folder to take the data from 
        Returns:
            none 
        """
        book = openpyxl.load_workbook(fileName)
        sheet = book.active

        for row in sheet.iter_rows(min_row = 2):
            
            moves = []
            for i in range(4,14,2):
                x = (row[i].value, row[i+1].value)
                moves.append(x)

            self.addCard(Card(row[0].value,row[1].value,row[2].value,row[3].value,moves))
            
    def addCard(self,theCard): 
        """
        adds a card to the deck, as long has it has all atributes. 

        Parameters:
            theCard- Card object to be passed into the deck 
        Returns:
            none 
        """
        try:
            if (isinstance(theCard, Card) and hasattr(theCard, 'name') and hasattr(theCard, 'type') and hasattr(theCard, 'hp') and hasattr(theCard, 'moves') and hasattr(theCard, 'shiny') and theCard.moves != None) : 
                self.deck.append(theCard)
                self.deckLength += 1
                self.totalAveDam += theCard.aveDam 
                if (theCard.shiny == 1): 
                    self.noOfShiny += 1
                
                print('Card ' + theCard.name + ' has been added to the deck ')
                print(theCard)
            else :
                raise Exception ('This Card does not have the nessisery attributes to add to the deck')

        except Exception as exp: 
            print('Error : ', exp, '\nCard ', theCard.name, 'has not been added to the deck')

    def rmCard(self, theCard): 
        """
        removes the card from a deck as long as it exists in the deck.  

        Parameters:
            theCard- Card object to be removed from deck 
        Returns:
            none 
        """
        try: 
            if (theCard in self.deck):
                self.deckLength -= 1 
                self.totalAveDam -= theCard.aveDam 
                if (theCard.shiny== 1): 
                    self.noOfShiny -= 1 
                self.deck.remove(theCard)
                
            else: 
                raise Exception ('Card is not in the deck so cannot be removed')
        except Exception as exp: 
            print('Error : ', exp)


    def getMostPowerful(self): 
        """
        gets most powerful card from the deck

        Parameters:
            none
        Returns:
            returns the most powerful card 
        """
        maxDam = max(c.aveDam for c in self.deck)
        mostPowerful = []
        for c in self.deck : 
            if c.aveDam == maxDam: 
                mostPowerful.append(c) 
        
        return mostPowerful

    def getAverageDamage(self): 
        """
        returns average damage of the deck 

        Parameters:
            none
        Returns:
            returns the average damage of the deck 
        """
        return round(self.totalAveDam/ self.deckLength , 1)

    def viewAllCards(self):
        """
        prints to screen all cards in the deck.

        Parameters:
            none
        Returns:
            none
        """
        self.viewCards(self.deck)

    def viewAllShinyCards(self): 
        """
        prints to screen all shiny cards in the deck.

        Parameters:
            none
        Returns:
            none
        """

        shinyDeck = [] 
        for c in self.deck: 
            if c.shiny == True:
                shinyDeck.append(c)

        self.viewCards(shinyDeck)

    def viewAllByType(self, theType): 
        """
        prints to screen all cards by type in the deck.

        Parameters:
            none
        Returns:
            none
        """
        typeDeck = [] 
        for c in self.deck: 
            if c.type == theType:
                typeDeck.append(c)

        self.viewCards(typeDeck)

    def getCards(self): 
        """
        returns the deck 

        Parameters:
            none
        Returns:
            returns the deck 
        """
        return self.deck  

    def saveToFile(self, fileName):
        """
        saves the deck to an xlsx file with the name filename

        Parameters:
            fileName - the file name and path if its not in the same folder to take the data from 
        Returns:
            none 
        """
        book2 = openpyxl.load_workbook(fileName)
        sheet2 = book2.active
        sheet2.append(('Name','Type','HP','Shiny','Move Name 1','Damage 1',	'Move Name 2','Damage 2','Move Name 3',	'Damage 3',	'Move Name 4',	'Damage 4',	'Move Name 5', 'Damage 5'))

        for c in self.deck: 
            attributes = []
            for i in (c.name, c.type, c.hp, c.shiny ) : 
                attributes.append(i)

            for key in c.moves : 
                attributes.append(key)
                attributes.append(c.moves[key])

            print(attributes)
            sheet2.append(attributes)
        
        book2.save(fileName)
            

    def viewCards(self, cardList):
        """
        functions to be called to view cards 

        Parameters:
            CardList - list of cards that want printing to screen 
        Returns:
            none 
        """
        for i in cardList: 
            print(i) 



def main(): 
    a = Card('tom','Magi', 300, 0, [('hit', 20),  ('kick', 50), ('slap', 10)])
    b = Card('mike','Magi', 600, [('hit', 53), ('kick', 50), ('slap', 70)], False)
    c = Card('timmothy dolton','Water', 300, [('hit', 1000), ('kick', 62), ('slap', 10)], True)
    print(a.name, a.type, a.hp, a.moves, a.shiny, a.totalDam, a.aveDam)
    print(a) 
    D = Deck() 
    D.addCard(a)
    D.addCard(b)
    D.addCard(c)
    D.viewCards(D.getMostPowerful())
    print(D.getAverageDamage())
    print(D.totalAveDam, D.deckLength)
    print(D)
    print('\nthis is the list of cards \n ')
    D.viewAllCards()
    print('\nthis is view all shiny cards : ')
    D.viewAllShinyCards()
    print('\nthis is viewAllByType')
    D.viewAllByType('Water')

    D.rmCard(a)
    print(D)
    print(len(D.deck))
    print(D.getAverageDamage())
    D.rmCard(Card('tom','Magi', 300, [('hit', 20),  ('kick', 50), ('slap', 10)], True))
    #D.addCard(c)
    #print(hasattr(a, 'name'))
    #print(hasattr(b, 'name'))
    #print(D.length)
    D = Deck()
    #D.imputFromFile('sampleDeck.xlsx')
    D.inputFromFile('sampleDeck.xlsx')
    a = Card('tom','Magi', 300, 0, [('hit', 20),  ('kick', 50), ('slap', 10)])
    D.addCard(a)
    D.saveToFile('practice.xlsx')





